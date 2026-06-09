#!/usr/bin/env python3
"""
Windows 事件日志安全分析工具
支持读取 Security / System / Application .evtx 日志文件
输出带可疑标注的 Excel 报告
"""

import os
import sys
import re
import glob
import argparse
import xml.etree.ElementTree as ET
from datetime import datetime, timezone

import Evtx.Evtx as evtx
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  可疑规则库
# ─────────────────────────────────────────────
SUSPICIOUS_RULES = {
    # 安全日志 (Security)
    "security": {
        # 登录失败
        4625: {"level": "HIGH",   "reason": "登录失败（暴力破解/凭据喷洒风险）"},
        # Kerberos 预认证失败
        4771: {"level": "HIGH",   "reason": "Kerberos 预认证失败（可能密码爆破）"},
        # 账户锁定
        4740: {"level": "HIGH",   "reason": "账户被锁定（多次登录失败触发）"},
        # 创建新账户
        4720: {"level": "HIGH",   "reason": "创建了新用户账户（可能植入后门账户）"},
        # 账户被启用
        4722: {"level": "MEDIUM", "reason": "用户账户被启用"},
        # 账户被添加到特权组
        4728: {"level": "HIGH",   "reason": "用户被加入安全组（可能权限提升）"},
        4732: {"level": "HIGH",   "reason": "用户被加入本地组（可能权限提升）"},
        4756: {"level": "HIGH",   "reason": "用户被加入通用组"},
        # 审计策略变更
        4719: {"level": "HIGH",   "reason": "系统审计策略被修改（攻击者常清除审计痕迹）"},
        # 日志被清除
        1102: {"level": "CRITICAL","reason": "安全日志被清除！（强烈可疑）"},
        # 特殊登录（管理员令牌）
        4672: {"level": "MEDIUM", "reason": "特权登录（分配了特殊权限）"},
        # 远程登录 (Type 10)
        4624: {"level": "INFO",   "reason": "成功登录（需关注 LogonType=10 远程桌面）"},
        # 密码策略变更
        4723: {"level": "MEDIUM", "reason": "用户尝试更改密码"},
        4724: {"level": "MEDIUM", "reason": "管理员重置了密码"},
        # 对象访问 – 进程创建
        4688: {"level": "MEDIUM", "reason": "新进程被创建（关注可疑命令行）"},
        # Windows Firewall 规则变更
        4946: {"level": "MEDIUM", "reason": "防火墙规则被添加"},
        4947: {"level": "MEDIUM", "reason": "防火墙规则被修改"},
        4950: {"level": "MEDIUM", "reason": "防火墙设置被修改"},
        # 计划任务
        4698: {"level": "HIGH",   "reason": "创建了计划任务（常用于持久化）"},
        4702: {"level": "MEDIUM", "reason": "计划任务被修改"},
        # 服务安装
        7045: {"level": "HIGH",   "reason": "安装了新服务（可能恶意软件持久化）"},
        # Pass-the-Hash 特征
        4648: {"level": "HIGH",   "reason": "使用显式凭据进行网络登录（Pass-the-Hash 特征）"},
        # Mimikatz/LSASS 访问
        4656: {"level": "MEDIUM", "reason": "请求了对象句柄（关注 LSASS 访问）"},
    },
    # 系统日志 (System)
    "system": {
        7045: {"level": "HIGH",   "reason": "安装了新服务（可能恶意软件持久化）"},
        7040: {"level": "MEDIUM", "reason": "服务启动类型被更改"},
        7036: {"level": "INFO",   "reason": "服务状态变更"},
        104:  {"level": "CRITICAL","reason": "系统日志被清除！（强烈可疑）"},
        # 意外关机/重启
        6008: {"level": "MEDIUM", "reason": "系统意外关机（可能被强制重启）"},
        # Windows 更新
        20:   {"level": "INFO",   "reason": "Windows Update 安装"},
        # 时间同步异常
        37:   {"level": "MEDIUM", "reason": "时钟同步异常"},
        # RDP 相关
        9009: {"level": "MEDIUM", "reason": "桌面窗口管理器异常（RDP 相关）"},
    },
    # 应用程序日志 (Application)
    "application": {
        # Windows Defender / AV
        1116: {"level": "CRITICAL","reason": "Defender 检测到恶意软件！"},
        1117: {"level": "CRITICAL","reason": "Defender 对恶意软件执行了操作"},
        1118: {"level": "HIGH",    "reason": "Defender 防病毒开始清除恶意软件"},
        1119: {"level": "HIGH",    "reason": "Defender 防病毒清除成功"},
        1120: {"level": "HIGH",    "reason": "Defender 防病毒清除失败"},
        # 应用程序崩溃
        1000: {"level": "MEDIUM",  "reason": "应用程序崩溃（关注系统进程崩溃）"},
        1001: {"level": "MEDIUM",  "reason": "Windows 错误报告"},
        # .NET Runtime 异常
        1026: {"level": "LOW",     "reason": ".NET Runtime 异常"},
        # MSSQL/DB 错误（SQL注入迹象）
        18456: {"level": "MEDIUM", "reason": "SQL Server 登录失败"},
    }
}

# 进程名中的可疑关键词
SUSPICIOUS_PROCESSES = [
    "powershell", "cmd.exe", "wscript", "cscript", "mshta", "regsvr32",
    "rundll32", "certutil", "bitsadmin", "wmic", "psexec", "net.exe",
    "net1.exe", "whoami", "mimikatz", "procdump", "at.exe", "schtasks",
    "vssadmin", "bcdedit", "wbadmin", "ntdsutil", "reg.exe", "regedit",
    "msiexec", "installutil", "regasm", "regsvcs", "msconfig", "netsh",
]

# 危险命令行关键词
SUSPICIOUS_CMDLINE = [
    "-enc", "-encodedcommand", "invoke-expression", "iex(",
    "downloadstring", "downloadfile", "webclient", "invoke-webrequest",
    "base64", "bypass", "-nop", "-windowstyle hidden",
    "vssadmin delete shadows", "bcdedit /set", "wbadmin delete",
    "net user /add", "net localgroup administrators",
    "reg add.*run", "schtasks /create",
]

LEVEL_COLOR = {
    "CRITICAL": "FF0000",
    "HIGH":     "FF6600",
    "MEDIUM":   "FFD700",
    "LOW":      "A9D18E",
    "INFO":     "BDD7EE",
}

LEVEL_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4, "NORMAL": 5}

NS = "{http://schemas.microsoft.com/win/2004/08/events/event}"


# ─────────────────────────────────────────────
#  解析单条事件
# ─────────────────────────────────────────────
def parse_event(record_xml: str, log_type: str) -> dict:
    try:
        root = ET.fromstring(record_xml)
    except ET.ParseError:
        return None

    sys_el  = root.find(f"{NS}System")
    if sys_el is None:
        return None

    def gtext(tag):
        el = sys_el.find(f"{NS}{tag}")
        return el.text if el is not None else ""

    def gattr(tag, attr):
        el = sys_el.find(f"{NS}{tag}")
        return el.get(attr, "") if el is not None else ""

    # 时间转本地（UTC+8）
    time_str = gattr("TimeCreated", "SystemTime")
    try:
        dt_utc = datetime.strptime(time_str[:26], "%Y-%m-%dT%H:%M:%S.%f")
        dt_utc = dt_utc.replace(tzinfo=timezone.utc)
        from datetime import timedelta
        dt_local = dt_utc.astimezone(timezone(timedelta(hours=8)))
        time_fmt = dt_local.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        time_fmt = time_str[:19] if time_str else ""

    event_id = int(gtext("EventID") or 0)
    level_code = gattr("Level", "")
    level_map = {"0":"Information","1":"Critical","2":"Error","3":"Warning","4":"Information","5":"Verbose"}
    level_name = level_map.get(gattr("Level",""), gtext("Level") or "")

    computer   = gtext("Computer")
    channel    = gattr("Channel", "") or gtext("Channel") or log_type
    provider   = gattr("Provider", "Name")

    # EventData / UserData – 提取所有 Data 字段
    data_pairs = {}
    ed = root.find(f"{NS}EventData")
    if ed is not None:
        for d in ed.findall(f"{NS}Data"):
            name  = d.get("Name", "")
            value = (d.text or "").strip()
            if name:
                data_pairs[name] = value
            # 无 Name 属性时拼接
            elif value:
                key = f"Data{len(data_pairs)}"
                data_pairs[key] = value

    # 构建描述摘要
    desc_parts = []
    for k, v in data_pairs.items():
        if v and v != "-":
            desc_parts.append(f"{k}={v}")
    description = " | ".join(desc_parts[:12])  # 最多12字段

    # 可疑检测
    suspicious_level = "NORMAL"
    suspicious_reason = ""

    rules = SUSPICIOUS_RULES.get(log_type.lower(), {})
    if event_id in rules:
        r = rules[event_id]
        suspicious_level  = r["level"]
        suspicious_reason = r["reason"]

    # 对 4624 补充检测：LogonType=10 才标高危
    if event_id == 4624 and log_type.lower() == "security":
        logon_type = data_pairs.get("LogonType", "")
        if logon_type == "10":
            suspicious_level  = "HIGH"
            suspicious_reason = "远程桌面(RDP)登录成功"
        elif logon_type == "3":
            suspicious_level  = "MEDIUM"
            suspicious_reason = "网络登录成功（Type 3）"
        elif logon_type in ("2",):
            suspicious_level  = "NORMAL"
            suspicious_reason = ""

    # 4688 进程创建：检测可疑进程/命令行
    if event_id == 4688:
        new_proc = data_pairs.get("NewProcessName", "").lower()
        cmdline  = data_pairs.get("CommandLine", "").lower()
        matched_proc = [p for p in SUSPICIOUS_PROCESSES if p in new_proc]
        matched_cmd  = [c for c in SUSPICIOUS_CMDLINE  if c in cmdline]
        if matched_proc or matched_cmd:
            suspicious_level  = "HIGH"
            suspicious_reason = "可疑进程/命令行: " + ", ".join(matched_proc + matched_cmd)

    # 日志清除
    if event_id in (1102, 104):
        suspicious_level  = "CRITICAL"
        suspicious_reason = SUSPICIOUS_RULES.get(log_type.lower(), {}).get(event_id, {}).get("reason", "日志被清除")

    return {
        "时间(UTC+8)":    time_fmt,
        "日志类型":       log_type.capitalize(),
        "事件ID":         event_id,
        "级别":           level_name,
        "计算机名":       computer,
        "来源/Provider":  provider,
        "可疑等级":       suspicious_level,
        "可疑原因":       suspicious_reason,
        "详细信息":       description,
    }


# ─────────────────────────────────────────────
#  解析 .evtx 文件
# ─────────────────────────────────────────────
def parse_evtx(filepath: str, log_type: str, max_records=50000) -> list:
    events = []
    count = 0
    try:
        with evtx.Evtx(filepath) as log:
            for record in log.records():
                try:
                    ev = parse_event(record.xml(), log_type)
                    if ev:
                        events.append(ev)
                except Exception:
                    pass
                count += 1
                if count >= max_records:
                    print(f"  [!] 已达到最大解析数 {max_records}，停止读取 {os.path.basename(filepath)}")
                    break
    except Exception as e:
        print(f"  [错误] 无法解析 {filepath}: {e}")
    print(f"  → {os.path.basename(filepath)}: 共解析 {len(events)} 条事件")
    return events


# ─────────────────────────────────────────────
#  写入 Excel
# ─────────────────────────────────────────────
HEADERS = ["时间(UTC+8)", "日志类型", "事件ID", "级别", "计算机名",
           "来源/Provider", "可疑等级", "可疑原因", "详细信息"]

COL_WIDTHS = [20, 10, 8, 12, 22, 28, 10, 38, 80]


def make_border(thin=True):
    s = Side(style="thin" if thin else "hair")
    return Border(left=s, right=s, top=s, bottom=s)


def write_excel(all_events: list, output_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. 汇总 Sheet ──────────────────────────
    ws_summary = wb.create_sheet("📊 汇总")
    write_summary(ws_summary, all_events)

    # ── 2. 可疑事件 Sheet ──────────────────────
    suspicious = [e for e in all_events if e["可疑等级"] != "NORMAL"]
    suspicious.sort(key=lambda x: (LEVEL_ORDER.get(x["可疑等级"], 9), x["时间(UTC+8)"]))
    ws_sus = wb.create_sheet("🚨 可疑事件")
    write_events_sheet(ws_sus, suspicious, title="可疑事件汇总（已按威胁等级排序）")

    # ── 3. 各类型日志 Sheet ────────────────────
    for log_type in ["Security", "System", "Application"]:
        evs = [e for e in all_events if e["日志类型"] == log_type]
        if not evs:
            continue
        icon = {"Security": "🔐", "System": "⚙️", "Application": "📋"}.get(log_type, "")
        ws = wb.create_sheet(f"{icon} {log_type}")
        write_events_sheet(ws, evs, title=f"{log_type} 日志（共 {len(evs)} 条）")

    # ── 4. 说明 Sheet ──────────────────────────
    ws_info = wb.create_sheet("ℹ️ 说明")
    write_info_sheet(ws_info)

    wb.save(output_path)
    print(f"\n✅ Excel 报告已保存：{output_path}")


def apply_header(ws, row=1):
    header_fill = PatternFill("solid", start_color="1F3864")
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = make_border()
    ws.row_dimensions[row].height = 22


def write_events_sheet(ws, events: list, title: str):
    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    title_cell.fill      = PatternFill("solid", start_color="2E4057")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    apply_header(ws, row=2)

    # 数据
    for ri, ev in enumerate(events, 3):
        level = ev.get("可疑等级", "NORMAL")
        row_color = LEVEL_COLOR.get(level, "FFFFFF") if level != "NORMAL" else None

        for ci, key in enumerate(HEADERS, 1):
            val  = ev.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=9,
                                  color="000000" if level in ("CRITICAL","HIGH") else "1A1A1A")
            cell.border    = make_border(thin=False)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == len(HEADERS)))

            if row_color:
                # 可疑行：整行淡色底 + 可疑等级列深色底
                fill_color = row_color + "30" if len(row_color) == 6 else row_color
                if key == "可疑等级":
                    cell.fill = PatternFill("solid", start_color=row_color)
                    cell.font = Font(bold=True, name="Arial", size=9,
                                     color="FFFFFF" if level in ("CRITICAL","HIGH","MEDIUM") else "000000")
                else:
                    bg = blend_color(row_color, alpha=0.18)
                    cell.fill = PatternFill("solid", start_color=bg)

        ws.row_dimensions[ri].height = 15

    # 列宽
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 冻结
    ws.freeze_panes = "A3"

    # 筛选
    if len(events) > 0:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{len(events)+2}"


def blend_color(hex_color: str, alpha=0.2) -> str:
    """将颜色与白色混合，生成淡色背景"""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r2 = int(r + (255 - r) * (1 - alpha))
    g2 = int(g + (255 - g) * (1 - alpha))
    b2 = int(b + (255 - b) * (1 - alpha))
    return f"{r2:02X}{g2:02X}{b2:02X}"


def write_summary(ws, all_events: list):
    from collections import Counter

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    # 主标题
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = "Windows 事件日志安全分析报告"
    c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=14)
    c.fill      = PatternFill("solid", start_color="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # 生成时间
    ws["A2"] = f"报告生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(italic=True, color="666666", size=9)
    ws.merge_cells("A2:D2")

    # 总体统计
    row = 4
    ws.cell(row=row, column=1, value="📋 总体统计").font = Font(bold=True, size=11, color="1F3864")
    row += 1

    stats = [
        ("总事件数", len(all_events)),
        ("可疑事件数", sum(1 for e in all_events if e["可疑等级"] != "NORMAL")),
        ("CRITICAL", sum(1 for e in all_events if e["可疑等级"] == "CRITICAL")),
        ("HIGH",     sum(1 for e in all_events if e["可疑等级"] == "HIGH")),
        ("MEDIUM",   sum(1 for e in all_events if e["可疑等级"] == "MEDIUM")),
        ("LOW",      sum(1 for e in all_events if e["可疑等级"] == "LOW")),
    ]
    for label, val in stats:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=val)
        c1.font = Font(name="Arial", size=10)
        c2.font = Font(bold=True, name="Arial", size=10)
        c2.alignment = Alignment(horizontal="center")
        level_color = {"CRITICAL":"FF0000","HIGH":"FF6600","MEDIUM":"FFD700","LOW":"A9D18E"}.get(label)
        if level_color:
            c1.fill = PatternFill("solid", start_color=blend_color(level_color, 0.25))
            c2.fill = PatternFill("solid", start_color=blend_color(level_color, 0.25))
        row += 1

    # 各日志类型统计
    row += 1
    ws.cell(row=row, column=1, value="📁 各日志类型统计").font = Font(bold=True, size=11, color="1F3864")
    row += 1

    type_counts = Counter(e["日志类型"] for e in all_events)
    for log_type, cnt in sorted(type_counts.items()):
        sus = sum(1 for e in all_events if e["日志类型"] == log_type and e["可疑等级"] != "NORMAL")
        ws.cell(row=row, column=1, value=log_type).font = Font(name="Arial", size=10)
        ws.cell(row=row, column=2, value=cnt).font      = Font(name="Arial", size=10)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=f"可疑: {sus}").font = Font(name="Arial", size=10,
                                                                       color="FF4444" if sus else "009900")
        row += 1

    # Top 可疑事件
    row += 1
    ws.cell(row=row, column=1, value="🚨 Top 可疑事件类型（按频次）").font = Font(bold=True, size=11, color="1F3864")
    row += 1

    sus_events = [e for e in all_events if e["可疑等级"] != "NORMAL"]
    reason_counts = Counter((e["事件ID"], e["可疑原因"]) for e in sus_events)

    for (eid, reason), cnt in reason_counts.most_common(20):
        ws.cell(row=row, column=1, value=f"EventID {eid}").font = Font(name="Arial", size=9, bold=True)
        ws.cell(row=row, column=2, value=cnt).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2).font = Font(name="Arial", size=9)
        ws.cell(row=row, column=3, value=reason).font = Font(name="Arial", size=9)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        row += 1


def write_info_sheet(ws):
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 70

    ws.merge_cells("A1:B1")
    ws["A1"].value     = "可疑等级说明 & 重点关注事件 ID"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    ws["A1"].fill      = PatternFill("solid", start_color="2E4057")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    rows = [
        ("等级",      "说明",                                           ""),
        ("CRITICAL",  "日志清除、Defender检测到恶意软件等最高危事件",    "FF0000"),
        ("HIGH",      "账户创建、暴力破解、权限提升、持久化行为等",       "FF6600"),
        ("MEDIUM",    "可疑进程启动、防火墙变更、账户操作等",             "FFD700"),
        ("LOW",       "一般告警，需结合上下文判断",                       "A9D18E"),
        ("INFO",      "信息类事件，正常登录等",                           "BDD7EE"),
        ("NORMAL",    "未匹配可疑规则，正常事件",                         "FFFFFF"),
        ("",          "",                                                ""),
        ("重点事件ID", "含义",                                           ""),
    ]

    event_desc = [
        ("4625", "登录失败 – 多次触发可能是暴力破解"),
        ("4624", "登录成功 – 重点关注 LogonType=10(RDP)"),
        ("4672", "特权登录 – 管理员令牌使用"),
        ("4720", "创建新账户 – 可能植入后门"),
        ("4728/4732", "加入特权组 – 权限提升"),
        ("4719", "审计策略变更 – 攻击者清除痕迹"),
        ("1102/104", "日志被清除 – 强烈可疑"),
        ("4688", "进程创建 – 关注 PowerShell/cmd 可疑命令行"),
        ("4698", "计划任务创建 – 常用持久化手段"),
        ("7045", "新服务安装 – 常用持久化手段"),
        ("4648", "显式凭据网络登录 – Pass-the-Hash 特征"),
        ("1116/1117", "Windows Defender 检测恶意软件"),
    ]

    row = 2
    for item in rows:
        label, desc, color = item
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=desc)
        if color and color != "FFFFFF" and label not in ("等级", "重点事件ID", ""):
            c1.fill = PatternFill("solid", start_color=color)
            c2.fill = PatternFill("solid", start_color=blend_color(color, 0.2))
        if label in ("等级", "重点事件ID"):
            c1.font = Font(bold=True, name="Arial", size=10)
            c2.font = Font(bold=True, name="Arial", size=10)
        else:
            c1.font = Font(name="Arial", size=9)
            c2.font = Font(name="Arial", size=9)
        row += 1

    for eid, desc in event_desc:
        ws.cell(row=row, column=1, value=eid).font  = Font(bold=True, name="Arial", size=9)
        ws.cell(row=row, column=2, value=desc).font = Font(name="Arial", size=9)
        row += 1


# ─────────────────────────────────────────────
#  主程序
# ─────────────────────────────────────────────
def detect_log_type(filepath: str) -> str:
    name = os.path.basename(filepath).lower()
    if "security" in name or "安全" in name:
        return "security"
    if "system" in name or "系统" in name:
        return "system"
    if "application" in name or "应用" in name:
        return "application"
    return "security"  # 默认


def main():
    parser = argparse.ArgumentParser(
        description="Windows 事件日志安全分析工具 – 输出带可疑标注的 Excel 报告",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog="""
示例:
  python analyze_logs.py Security.evtx System.evtx Application.evtx
  python analyze_logs.py *.evtx -o 安全分析报告.xlsx
  python analyze_logs.py C:\\logs\\*.evtx --max 100000
        """
    )
    parser.add_argument("logs", nargs="*",
                        help="一个或多个 .evtx 文件路径（支持通配符）")
    parser.add_argument("-o", "--output", default="Windows日志安全分析报告.xlsx",
                        help="输出 Excel 文件名（默认：Windows日志安全分析报告.xlsx）")
    parser.add_argument("--max", type=int, default=50000,
                        help="每个日志文件最大解析条数（默认50000）")
    args = parser.parse_args()

    # 展开通配符
    files = []
    for pattern in args.logs:
        matched = glob.glob(pattern)
        if matched:
            files.extend(matched)
        elif os.path.isfile(pattern):
            files.append(pattern)

    if not files:
        # 自动搜索当前目录
        files = glob.glob("*.evtx") + glob.glob("**/*.evtx", recursive=True)

    if not files:
        print("❌ 未找到 .evtx 文件。请指定日志文件路径，例如：")
        print("   python analyze_logs.py Security.evtx System.evtx Application.evtx")
        sys.exit(1)

    print(f"📂 找到 {len(files)} 个日志文件：")
    for f in files:
        print(f"   {f}")
    print()

    all_events = []
    for fp in files:
        log_type = detect_log_type(fp)
        print(f"⏳ 解析 {os.path.basename(fp)} [{log_type}] ...")
        evs = parse_evtx(fp, log_type, max_records=args.max)
        all_events.extend(evs)

    print(f"\n📊 共解析事件：{len(all_events)} 条")
    sus = [e for e in all_events if e["可疑等级"] != "NORMAL"]
    print(f"🚨 可疑事件：{len(sus)} 条")

    crit = [e for e in sus if e["可疑等级"] == "CRITICAL"]
    if crit:
        print(f"\n⚠️  CRITICAL 级别事件（{len(crit)} 条）：")
        for e in crit[:10]:
            print(f"   [{e['时间(UTC+8)']}] EventID {e['事件ID']} – {e['可疑原因']}")

    write_excel(all_events, args.output)


if __name__ == "__main__":
    main()
